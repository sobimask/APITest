##本文件是读取接口文档xlsx，获取每一行的值

from openpyxl import load_workbook
from common.log import log
from common.readconf import apifile_dir

log('开始读取api文档')
class ReadCase():
    def get_rows_data(self,case_file,case_sheet):
        log('创建空datalist')
        datalist = []
        try:
            if datalist != None:
                datalist.clear()
            else:
                datalist = []
        finally:
            workbook = load_workbook(case_file)
            sheets = workbook[case_sheet]
            rows = sheets.iter_rows()
            log('获取工作表格{workbook}'.format(workbook=workbook))
            log('获取sheets为{sheets}'.format(sheets=sheets))
            for item in rows:
                if item[0].value == 'url':
                    continue;
                list = []
                for col in item:
                    if item[0].value ==None:
                        continue;
                    list.append(col.value)
                # 用例读取到的列表值
                #优化读取case，不会读取到空白行添加到列表
                if list !=[]:
                    datalist.append(list)
            log('读取到列表{datalist}'.format(datalist=datalist))
            return datalist

# ReadCase().get_rows_data(apifile_dir+'DeviceMapController.xlsx','Sheet1')
# ReadCase().get_rows_data(apifile_dir+'dashboard.xlsx','Sheet')