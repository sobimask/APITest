from openpyxl import load_workbook


def write_data(filename, sheetname, row, column, data):
    """
    :param row: 指定在某一行写
    :param column: 指定在某一列写
    :param data: 待写入的数据
    """
    wb = load_workbook(filename)
    ws = wb[sheetname]
    ws.cell(row, column, value=data)
    wb.save(filename)
